"""Streamlit-independent report generation for calculated manual workflows."""

from __future__ import annotations

import csv
import json
from datetime import datetime, timezone
from io import BytesIO, StringIO
from typing import Any

from openpyxl import Workbook

from hcmcalc.ui.manual_freeway import (
    MANUAL_FREEWAY_LIMITATIONS,
    freeway_display_outputs,
)
from hcmcalc.ui.manual_multilane import (
    MANUAL_MULTILANE_LIMITATIONS,
    multilane_display_outputs,
)
from hcmcalc.ui.units import MILES_TO_KILOMETERS, display_outputs


REPORT_SCHEMA_VERSION = "0.1"
SUPPORTED_CALCULATION_TYPES = {
    "manual_single_segment",
    "manual_facility_v0",
    "manual_multilane_v0",
    "manual_basic_freeway_v0",
}
SUPPORTED_EXPORT_FORMATS = {"csv", "xlsx", "markdown", "json"}

SINGLE_SEGMENT_LIMITATIONS = [
    "This is a limited calculator workflow.",
    "Unsupported combinations remain guarded.",
    "This is not a full general HCM Chapter 15 calculator.",
]
FACILITY_LIMITATIONS = [
    "Manual Facility Calculator v0.1 is template-backed.",
    "Facility support is anchored to validated Example 3/4 behavior.",
    "This is not full general facility support.",
    "Unsupported combinations remain guarded.",
]


class ReportingError(ValueError):
    """Raised when a report cannot be generated safely."""


def build_report(
    calculation_type: str,
    result: dict[str, Any] | None,
    unit_system: str,
    *,
    inputs: dict[str, Any] | list[dict[str, Any]] | None = None,
    audit_record: dict[str, Any] | None = None,
    template_id: str | None = None,
    generated_at: str | None = None,
) -> dict[str, Any]:
    """Build a report-friendly structure from an existing calculated result."""

    _validate_result(calculation_type, result)
    unit_system = _normalize_unit_system(unit_system)
    assert result is not None
    _ensure_json_serializable(
        {"result": result, "inputs": inputs, "audit_record": audit_record}
    )
    outputs = result["outputs"]
    timestamp = generated_at or _audit_timestamp(audit_record) or datetime.now(
        timezone.utc
    ).isoformat()

    if calculation_type == "manual_single_segment":
        report = _single_segment_report(
            result, outputs, unit_system, inputs, audit_record, timestamp
        )
    elif calculation_type == "manual_facility_v0":
        report = _facility_report(
            result,
            outputs,
            unit_system,
            inputs,
            audit_record,
            template_id,
            timestamp,
        )
    elif calculation_type == "manual_multilane_v0":
        report = _multilane_report(
            result,
            outputs,
            unit_system,
            inputs,
            audit_record,
            template_id,
            timestamp,
        )
    else:
        report = _freeway_report(
            result,
            outputs,
            unit_system,
            inputs,
            audit_record,
            template_id,
            timestamp,
        )
    _ensure_json_serializable(report)
    return report


def export_report(report: dict[str, Any], export_format: str) -> str | bytes:
    """Render an existing report structure without recalculating results."""

    export_format = str(export_format).strip().lower()
    if export_format not in SUPPORTED_EXPORT_FORMATS:
        raise ReportingError(f"Unsupported export format: {export_format}.")
    _ensure_json_serializable(report)
    if export_format == "json":
        return json.dumps(report, indent=2)
    if export_format == "markdown":
        return report_to_markdown(report)
    if export_format == "csv":
        return report_to_csv(report)
    return report_to_xlsx(report)


def report_filename(report: dict[str, Any], extension: str) -> str:
    """Return the required timestamped filename for a report."""

    extension = str(extension).lower().lstrip(".")
    if extension not in {"csv", "xlsx", "md", "json"}:
        raise ReportingError(f"Unsupported export format: {extension}.")
    calculation_type = report.get("calculation_type")
    if calculation_type not in SUPPORTED_CALCULATION_TYPES:
        raise ReportingError(f"Unsupported calculation_type: {calculation_type}.")
    try:
        timestamp = datetime.fromisoformat(str(report["generated_at"]).replace("Z", "+00:00"))
    except (KeyError, ValueError) as exc:
        raise ReportingError("Report generated_at must be an ISO-8601 timestamp.") from exc
    workflow = {
        "manual_single_segment": "single_segment",
        "manual_facility_v0": "facility",
        "manual_multilane_v0": "multilane",
        "manual_basic_freeway_v0": "basic_freeway",
    }[calculation_type]
    return f"hcm_ch15_{workflow}_report_{timestamp:%Y%m%d_%H%M%S}.{extension}"


def report_to_csv(report: dict[str, Any]) -> str:
    """Render report sections into a copy-ready CSV document."""

    output = StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow([report["title"]])
    for key in ("report_type", "calculation_type", "unit_system", "generated_at"):
        writer.writerow([_label(key), report.get(key, "")])
    _write_key_value_csv(writer, "Summary", report["results_summary"])
    _write_key_value_csv(writer, "Inputs", report["inputs_summary"])
    if report.get("normalized_engine_inputs_summary"):
        _write_key_value_csv(
            writer,
            "Normalized Engine Inputs",
            report["normalized_engine_inputs_summary"],
        )
    _write_table_csv(writer, "Segment Results", report["segment_results"])
    _write_list_csv(writer, "Assumptions", report["assumptions"])
    _write_list_csv(writer, "Warnings", report["warnings"])
    _write_list_csv(writer, "Limitations", report["limitations"])
    _write_key_value_csv(writer, "Audit", report["audit_summary"])
    if report.get("intermediate_values"):
        _write_table_csv(writer, "Intermediate Values", report["intermediate_values"])
    return output.getvalue()


def report_to_markdown(report: dict[str, Any]) -> str:
    """Render a clean copy-ready Markdown report."""

    lines = [
        f"# {report['title']}",
        "",
        f"- **Calculation type:** {report['calculation_type']}",
        f"- **Unit system:** {report['unit_system']}",
        f"- **Generated at:** {report['generated_at']}",
        "",
        "## Summary Result",
        "",
        *_markdown_key_value_table(report["results_summary"]),
        "",
        "## Key Inputs",
        "",
        *_markdown_key_value_table(report["inputs_summary"]),
    ]
    if report.get("normalized_engine_inputs_summary"):
        lines.extend(
            [
                "",
                "## Normalized Engine Inputs",
                "",
                *_markdown_key_value_table(report["normalized_engine_inputs_summary"]),
            ]
        )
    if report["segment_results"]:
        lines.extend(["", "## Segment Results", "", *_markdown_table(report["segment_results"])])
    for heading, key in (
        ("Assumptions", "assumptions"),
        ("Warnings", "warnings"),
        ("Limitations", "limitations"),
    ):
        values = report[key] or [f"No {heading.lower()} reported."]
        lines.extend(["", f"## {heading}", "", *(f"- {value}" for value in values)])
    lines.extend(
        [
            "",
            "## Audit / Source References Summary",
            "",
            *_markdown_key_value_table(report["audit_summary"]),
        ]
    )
    if report.get("intermediate_values"):
        lines.extend(
            [
                "",
                "## Intermediate Values",
                "",
                *_markdown_table(report["intermediate_values"]),
            ]
        )
    lines.append("")
    return "\n".join(lines)


def report_to_xlsx(report: dict[str, Any]) -> bytes:
    """Render a lightweight report workbook."""

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    _append_metadata(summary, report)
    _append_key_values(summary, report["results_summary"])

    inputs = workbook.create_sheet("Inputs")
    _append_key_values(inputs, report["inputs_summary"])
    if report.get("normalized_engine_inputs_summary"):
        inputs.append([])
        inputs.append(["Normalized Engine Inputs"])
        _append_key_values(inputs, report["normalized_engine_inputs_summary"])

    segments = workbook.create_sheet("Segment Results")
    _append_table(segments, report["segment_results"])

    context = workbook.create_sheet("Assumptions Warnings Limits")
    context.append(["Category", "Text"])
    for key in ("assumptions", "warnings", "limitations"):
        for value in report[key]:
            context.append([_label(key), value])

    audit = workbook.create_sheet("Audit")
    _append_key_values(audit, report["audit_summary"])
    if report.get("intermediate_values"):
        audit.append([])
        audit.append(["Intermediate Values"])
        _append_table(audit, report["intermediate_values"])

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.column_dimensions["A"].width = 38
        worksheet.column_dimensions["B"].width = 70
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _single_segment_report(
    result: dict[str, Any],
    outputs: dict[str, Any],
    unit_system: str,
    inputs: Any,
    audit_record: dict[str, Any] | None,
    timestamp: str,
) -> dict[str, Any]:
    metrics = display_outputs(outputs, unit_system)
    summary = [
        {"label": "Level of service", "value": outputs["level_of_service"], "unit": None},
        *[
            {"label": metric["label"], "value": metric["value"], "unit": metric["unit"]}
            for metric in metrics.values()
        ],
    ]
    return _base_report(
        title="HCM7 Chapter 15 Manual Single Segment Report",
        report_type="HCM Chapter 15 calculation report",
        calculation_type="manual_single_segment",
        unit_system=unit_system,
        timestamp=timestamp,
        inputs=_input_records(inputs or {}, unit_system),
        results=summary,
        segment_results=[_single_segment_row(outputs, unit_system)],
        result=result,
        audit_record=audit_record,
        limitations=SINGLE_SEGMENT_LIMITATIONS,
    )


def _facility_report(
    result: dict[str, Any],
    outputs: dict[str, Any],
    unit_system: str,
    inputs: Any,
    audit_record: dict[str, Any] | None,
    template_id: str | None,
    timestamp: str,
) -> dict[str, Any]:
    metric = unit_system == "metric"
    summary = [
        {"label": "Facility level of service", "value": outputs["facility_level_of_service"], "unit": None},
        {"label": "Facility follower density", "value": outputs["facility_follower_density_followers_mi_ln"] / MILES_TO_KILOMETERS if metric else outputs["facility_follower_density_followers_mi_ln"], "unit": "fol/km/ln" if metric else "fol/mi/ln"},
        {"label": "Facility average speed", "value": outputs["facility_average_speed_mph"] * MILES_TO_KILOMETERS if metric else outputs["facility_average_speed_mph"], "unit": "km/h" if metric else "mph"},
        {"label": "Facility length", "value": outputs["facility_length_mi"] * MILES_TO_KILOMETERS if metric else outputs["facility_length_mi"], "unit": "km" if metric else "mi"},
        {"label": "Segment count", "value": len(outputs["segments"]), "unit": None},
    ]
    input_records = _input_records(inputs or [], unit_system)
    if template_id:
        input_records.insert(0, {"label": "Template ID", "value": template_id, "unit": None})
    return _base_report(
        title="HCM7 Chapter 15 Manual Facility Calculator v0.1 Report",
        report_type="HCM Chapter 15 limited template-backed facility report",
        calculation_type="manual_facility_v0",
        unit_system=unit_system,
        timestamp=timestamp,
        inputs=input_records,
        results=summary,
        segment_results=[_facility_segment_row(row, unit_system) for row in outputs["segments"]],
        result=result,
        audit_record=audit_record,
        limitations=FACILITY_LIMITATIONS,
    )


def _multilane_report(
    result: dict[str, Any],
    outputs: dict[str, Any],
    unit_system: str,
    inputs: Any,
    audit_record: dict[str, Any] | None,
    template_id: str | None,
    timestamp: str,
) -> dict[str, Any]:
    display = multilane_display_outputs(outputs, unit_system)
    summary = [
        {"label": "Level of service", "value": outputs["level_of_service"], "unit": None},
        *[
            {"label": _label(name), "value": metric["value"], "unit": metric["unit"]}
            for name, metric in display.items()
        ],
        {
            "label": "Heavy vehicle adjustment factor",
            "value": outputs["heavy_vehicle_adjustment_factor"],
            "unit": None,
        },
        {"label": "Capacity check", "value": outputs["capacity_check"], "unit": None},
    ]
    input_records = _multilane_input_records(inputs or {}, unit_system)
    if template_id:
        input_records.insert(
            0, {"label": "Validated template", "value": template_id, "unit": None}
        )
    report = _base_report(
        title="HCM7 Manual Multilane Highway Segment v0.1 Report",
        report_type="HCM Manual Multilane validated-path calculation report",
        calculation_type="manual_multilane_v0",
        unit_system=unit_system,
        timestamp=timestamp,
        inputs=input_records,
        results=summary,
        segment_results=[],
        result=result,
        audit_record=audit_record,
        limitations=MANUAL_MULTILANE_LIMITATIONS,
    )
    report["selected_validated_template"] = template_id
    report["normalized_engine_inputs_summary"] = _input_records(
        (audit_record or {}).get("submitted_inputs", {}), "imperial"
    )
    return report


def _freeway_report(
    result: dict[str, Any],
    outputs: dict[str, Any],
    unit_system: str,
    inputs: Any,
    audit_record: dict[str, Any] | None,
    preset_id: str | None,
    timestamp: str,
) -> dict[str, Any]:
    display = freeway_display_outputs(outputs, unit_system)
    summary = [
        {"label": "Level of service", "value": outputs["level_of_service"], "unit": None},
        *[
            {"label": _label(name), "value": metric["value"], "unit": metric["unit"]}
            for name, metric in display.items()
        ],
        {
            "label": "Heavy vehicle adjustment factor",
            "value": outputs["heavy_vehicle_adjustment_factor"],
            "unit": None,
        },
        {"label": "Capacity check", "value": outputs["capacity_check"], "unit": None},
    ]
    input_records = _freeway_input_records(inputs or {}, unit_system)
    if preset_id:
        input_records.insert(
            0, {"label": "Validated preset", "value": preset_id, "unit": None}
        )
    report = _base_report(
        title="Manual Basic Freeway Segment Calculator Report",
        report_type="HCM Manual Basic Freeway Segment validated-path calculation report",
        calculation_type="manual_basic_freeway_v0",
        unit_system=unit_system,
        timestamp=timestamp,
        inputs=input_records,
        results=summary,
        segment_results=[],
        result=result,
        audit_record=audit_record,
        limitations=MANUAL_FREEWAY_LIMITATIONS,
    )
    report["selected_validated_preset"] = preset_id
    report["support_scope"] = "Basic Freeway Segment only; BF-CH26-001-compatible validated path."
    report["normalized_engine_inputs_summary"] = _input_records(
        (audit_record or {}).get("submitted_inputs", {}), "imperial"
    )
    return report


def _multilane_input_records(
    inputs: Any, unit_system: str
) -> list[dict[str, Any]]:
    if not isinstance(inputs, dict):
        raise ReportingError("Manual Multilane report inputs must be an object.")
    metric = unit_system == "metric"
    units = {
        "segment_length": "km" if metric else "ft",
        "posted_speed_limit": "km/h" if metric else "mph",
        "lane_width": "m" if metric else "ft",
        "roadside_lateral_clearance": "m" if metric else "ft",
        "access_point_density": "access points/km" if metric else "access points/mi",
        "demand_volume_veh_h": "veh/h",
        "heavy_vehicle_percent": "%",
        "grade_percent": "%",
    }
    return [
        {"label": _label(key), "value": value, "unit": units.get(key)}
        for key, value in inputs.items()
    ]


def _freeway_input_records(inputs: Any, unit_system: str) -> list[dict[str, Any]]:
    if not isinstance(inputs, dict):
        raise ReportingError("Manual Basic Freeway report inputs must be an object.")
    metric = unit_system == "metric"
    units = {
        "segment_length": "km" if metric else "mi",
        "free_flow_speed": "km/h" if metric else "mph",
        "base_free_flow_speed": "km/h" if metric else "mph",
        "lane_width": "m" if metric else "ft",
        "right_side_lateral_clearance": "m" if metric else "ft",
        "total_ramp_density": "ramps/km" if metric else "ramps/mi",
        "demand_volume_veh_h": "veh/h",
        "heavy_vehicle_percent": "%",
    }
    return [
        {"label": _label(key), "value": value, "unit": units.get(key)}
        for key, value in inputs.items()
    ]


def _base_report(**values: Any) -> dict[str, Any]:
    result = values.pop("result")
    audit_record = values.pop("audit_record")
    return {
        "schema_version": REPORT_SCHEMA_VERSION,
        "title": values["title"],
        "report_type": values["report_type"],
        "calculation_type": values["calculation_type"],
        "unit_system": values["unit_system"],
        "inputs_summary": values["inputs"],
        "results_summary": values["results"],
        "segment_results": values["segment_results"],
        "assumptions": list(result.get("assumptions", [])),
        "warnings": list(result.get("warnings", [])),
        "limitations": list(values["limitations"]),
        "audit_summary": _audit_summary(result, audit_record),
        "intermediate_values": list(result.get("intermediate_values", [])),
        "generated_at": values["timestamp"],
    }


def _single_segment_row(outputs: dict[str, Any], unit_system: str) -> dict[str, Any]:
    metric = unit_system == "metric"
    return {
        "Segment type": outputs.get("segment_type"),
        f"Length ({'km' if metric else 'mi'})": _convert(outputs.get("segment_length_mi"), MILES_TO_KILOMETERS if metric else 1),
        f"Average speed ({'km/h' if metric else 'mph'})": _convert(outputs.get("average_speed_mph"), MILES_TO_KILOMETERS if metric else 1),
        "Percent followers (%)": outputs.get("percent_followers"),
        f"Follower density ({'fol/km/ln' if metric else 'fol/mi/ln'})": _convert(outputs.get("follower_density_followers_mi_ln"), 1 / MILES_TO_KILOMETERS if metric else 1),
        "Level of service": outputs.get("level_of_service"),
    }


def _facility_segment_row(segment: dict[str, Any], unit_system: str) -> dict[str, Any]:
    row = _single_segment_row(segment, unit_system)
    return {"Segment ID": segment.get("segment_id"), **row}


def _input_records(inputs: Any, unit_system: str) -> list[dict[str, Any]]:
    if isinstance(inputs, list):
        return [
            {"label": f"Segment {index} {label}", "value": value, "unit": unit}
            for index, row in enumerate(inputs, start=1)
            for label, value, unit in _input_items(row, unit_system)
        ]
    if not isinstance(inputs, dict):
        raise ReportingError("Report inputs must be an object or a list of segment objects.")
    return [
        {"label": label, "value": value, "unit": unit}
        for label, value, unit in _input_items(inputs, unit_system)
    ]


def _input_items(inputs: dict[str, Any], unit_system: str) -> list[tuple[str, Any, str | None]]:
    if not isinstance(inputs, dict):
        raise ReportingError("Each report input row must be an object.")
    units = {
        "segment_length": "km" if unit_system == "metric" else "mi",
        "posted_speed": "km/h" if unit_system == "metric" else "mph",
        "lane_width": "m" if unit_system == "metric" else "ft",
        "shoulder_width": "m" if unit_system == "metric" else "ft",
        "access_point_density": "access points/km" if unit_system == "metric" else "access points/mi",
        "analysis_direction_volume": "veh/h",
        "analysis_direction_volume_veh_h": "veh/h",
        "opposing_direction_volume": "veh/h",
        "opposing_direction_volume_veh_h": "veh/h",
        "heavy_vehicle_percent": "%",
        "grade_percent": "%",
        "segment_length_ft": "ft (engine-native Imperial)",
        "posted_speed_limit_mph": "mph (engine-native Imperial)",
        "lane_width_ft": "ft (engine-native Imperial)",
        "roadside_lateral_clearance_ft": "ft (engine-native Imperial)",
        "access_point_density_per_mi": "access points/mi (engine-native Imperial)",
        "segment_length_mi": "mi (engine-native Imperial)",
        "free_flow_speed_mph": "mph (engine-native Imperial)",
        "base_free_flow_speed_mph": "mph (engine-native Imperial)",
        "right_side_lateral_clearance_ft": "ft (engine-native Imperial)",
        "total_ramp_density_per_mi": "ramps/mi (engine-native Imperial)",
        "demand_volume_veh_h": "veh/h",
    }
    return [(_label(key), value, units.get(key)) for key, value in inputs.items()]


def _audit_summary(result: dict[str, Any], audit_record: dict[str, Any] | None) -> list[dict[str, Any]]:
    records = [
        {"label": "Method", "value": result.get("method"), "unit": None},
        {"label": "Facility type", "value": result.get("facility_type"), "unit": None},
    ]
    for path, value in _flatten(audit_record or {}):
        lowered = path.lower()
        if any(
            term in lowered
            for term in (
                "source",
                "reference",
                "validation_basis",
                "scope_status",
                "support_status",
                "template_id",
            )
        ):
            records.append({"label": path, "value": value, "unit": None})
    return records


def _flatten(value: Any, prefix: str = "") -> list[tuple[str, Any]]:
    if isinstance(value, dict):
        return [
            item
            for key, child in value.items()
            for item in _flatten(child, f"{prefix}.{key}" if prefix else str(key))
        ]
    if isinstance(value, list):
        return [(prefix, "; ".join(str(item) for item in value))]
    return [(prefix, value)]


def _write_key_value_csv(writer: Any, heading: str, records: list[dict[str, Any]]) -> None:
    writer.writerow([])
    writer.writerow([heading])
    writer.writerow(["Label", "Value", "Unit"])
    for record in records:
        writer.writerow([record["label"], _cell(record.get("value")), record.get("unit") or ""])


def _write_table_csv(writer: Any, heading: str, rows: list[dict[str, Any]]) -> None:
    writer.writerow([])
    writer.writerow([heading])
    if not rows:
        writer.writerow(["No rows"])
        return
    writer.writerow(list(rows[0]))
    for row in rows:
        writer.writerow([_cell(value) for value in row.values()])


def _write_list_csv(writer: Any, heading: str, values: list[Any]) -> None:
    writer.writerow([])
    writer.writerow([heading])
    writer.writerow(["Text"])
    for value in values:
        writer.writerow([_cell(value)])


def _markdown_key_value_table(records: list[dict[str, Any]]) -> list[str]:
    rows = [{"Item": record["label"], "Value": record.get("value"), "Unit": record.get("unit") or ""} for record in records]
    return _markdown_table(rows)


def _markdown_table(rows: list[dict[str, Any]]) -> list[str]:
    if not rows:
        return ["No rows reported."]
    headers = list(rows[0])
    return [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
        *["| " + " | ".join(_markdown_cell(row.get(header)) for header in headers) + " |" for row in rows],
    ]


def _append_metadata(worksheet: Any, report: dict[str, Any]) -> None:
    for key in ("title", "report_type", "calculation_type", "unit_system", "generated_at"):
        worksheet.append([_label(key), report[key]])
    worksheet.append([])


def _append_key_values(worksheet: Any, records: list[dict[str, Any]]) -> None:
    worksheet.append(["Label", "Value", "Unit"])
    for record in records:
        worksheet.append([record["label"], _cell(record.get("value")), record.get("unit")])


def _append_table(worksheet: Any, rows: list[dict[str, Any]]) -> None:
    if not rows:
        worksheet.append(["No rows"])
        return
    worksheet.append(list(rows[0]))
    for row in rows:
        worksheet.append([_cell(value) for value in row.values()])


def _validate_result(calculation_type: str, result: dict[str, Any] | None) -> None:
    if result is None:
        raise ReportingError("A calculated result is required.")
    if calculation_type not in SUPPORTED_CALCULATION_TYPES:
        raise ReportingError(f"Unsupported calculation_type: {calculation_type}.")
    if not isinstance(result, dict) or not isinstance(result.get("outputs"), dict):
        raise ReportingError("Malformed result object: outputs must be an object.")
    required = (
        {
            "level_of_service",
            "follower_density_followers_mi_ln",
            "average_speed_mph",
            "percent_followers",
            "demand_flow_rate_veh_h",
            "capacity_veh_h",
            "free_flow_speed_mph",
        }
        if calculation_type == "manual_single_segment"
        else (
            {
                "facility_level_of_service",
                "facility_follower_density_followers_mi_ln",
                "facility_average_speed_mph",
                "facility_length_mi",
                "segments",
            }
            if calculation_type == "manual_facility_v0"
            else (
                {
                    "density_pc_mi_ln",
                    "level_of_service",
                    "speed_used_for_density_mph",
                    "demand_flow_rate_pc_h_ln",
                    "adjusted_free_flow_speed_mph",
                    "base_free_flow_speed_mph",
                    "heavy_vehicle_adjustment_factor",
                    "capacity_pc_h_ln",
                    "capacity_check",
                }
                if calculation_type == "manual_multilane_v0"
                else {
                "density_pc_mi_ln",
                "level_of_service",
                "speed_used_for_density_mph",
                "demand_flow_rate_pc_h_ln",
                "adjusted_free_flow_speed_mph",
                "base_free_flow_speed_mph",
                "heavy_vehicle_adjustment_factor",
                "capacity_pc_h_ln",
                "adjusted_capacity_pc_h_ln",
                "capacity_check",
            }
            )
        )
    )
    if not required.issubset(result["outputs"]):
        raise ReportingError("Malformed result object: required outputs are missing.")
    if calculation_type == "manual_facility_v0" and not isinstance(result["outputs"]["segments"], list):
        raise ReportingError("Malformed result object: facility segments must be a list.")
    if calculation_type == "manual_facility_v0" and any(
        not isinstance(segment, dict) for segment in result["outputs"]["segments"]
    ):
        raise ReportingError("Malformed result object: each facility segment must be an object.")


def _normalize_unit_system(unit_system: str) -> str:
    normalized = str(unit_system).strip().lower()
    if normalized not in {"metric", "imperial"}:
        raise ReportingError("unit_system must be metric or imperial.")
    return normalized


def _audit_timestamp(audit_record: dict[str, Any] | None) -> str | None:
    return audit_record.get("generated_at") if isinstance(audit_record, dict) else None


def _ensure_json_serializable(value: Any) -> None:
    try:
        json.dumps(value, allow_nan=False)
    except (TypeError, ValueError) as exc:
        raise ReportingError(f"Non-serializable report field: {exc}") from exc


def _convert(value: Any, factor: float) -> Any:
    return float(value) * factor if value is not None else None


def _cell(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, sort_keys=True)
    return value


def _markdown_cell(value: Any) -> str:
    return str(_cell(value) if value is not None else "").replace("|", r"\|").replace("\n", " ")


def _label(value: str) -> str:
    return str(value).replace("_", " ").strip().title()
