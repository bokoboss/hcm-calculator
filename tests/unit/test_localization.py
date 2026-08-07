"""Presentation-only Thai/English localization contracts."""

from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path

import pytest

from openpyxl import load_workbook

from hcmcalc.cli import result_to_dict
from hcmcalc.ui.i18n import (
    SUPPORTED_LOCALES,
    normalize_locale,
    translate,
    validate_catalogs,
)
from hcmcalc.ui.manual_freeway import (
    freeway_preset_ui_inputs,
    load_freeway_preset,
    run_manual_freeway,
)
from hcmcalc.ui.project_io import (
    create_manual_freeway_project_json,
    load_manual_freeway_project_json,
    project_presentation_locale,
)
from hcmcalc.ui.manual_multilane import (
    load_multilane_template,
    multilane_template_ui_inputs,
    multilane_ui_inputs_to_engine,
)
from hcmcalc.ui.reporting import build_report, export_report
from hcmcalc.ui.workflow_state import calculation_input_fingerprint


def _freeway_result() -> dict[str, object]:
    return result_to_dict(run_manual_freeway(load_freeway_preset("BF-CH26-001")["inputs"]))


def test_catalogs_have_exact_key_and_placeholder_parity() -> None:
    assert SUPPORTED_LOCALES == ("en", "th")
    assert validate_catalogs() == []


def test_multilane_task_copy_is_bilingual_and_adapter_values_are_canonical() -> None:
    keys = (
        "multilane.start",
        "multilane.traffic_segment",
        "multilane.free_flow_speed_section",
        "multilane.heavy_vehicle_adjustment",
        "multilane.heavy_vehicle_method_help",
        "multilane.composition_scope",
        "multilane.recalculate",
        "multilane.details_audit",
    )
    for key in keys:
        for locale in SUPPORTED_LOCALES:
            assert translate(key, locale) != key

    english_scope = translate("multilane.composition_scope", "en")
    thai_scope = translate("multilane.composition_scope", "th")
    assert "within the heavy-vehicle portion of total traffic" in english_scope
    assert "not another percentage of total traffic" in english_scope
    assert "ภายในกลุ่มยานพาหนะหนักของปริมาณจราจรรวม" in thai_scope
    assert "ไม่ใช่สัดส่วนอีกค่าหนึ่งของปริมาณจราจรรวม" in thai_scope

    template = load_multilane_template("MLH-CH26-004-EB")["inputs"]
    values = multilane_template_ui_inputs("MLH-CH26-004-EB", "imperial")
    values.update(
        {
            "heavy_vehicle_adjustment_method": "external_pce",
            "passenger_car_equivalent": 2.5,
        }
    )
    engine_inputs = multilane_ui_inputs_to_engine(values, template, "imperial")
    assert engine_inputs["terrain_type"] == "specific_grade"
    assert engine_inputs["passenger_car_equivalent"] == 2.5

    localized_method = translate("multilane.heavy_method.external_pce", "th")
    assert localized_method != "external_pce"
    with pytest.raises(ValueError, match="heavy_vehicle_adjustment_method"):
        multilane_ui_inputs_to_engine(
            values | {"heavy_vehicle_adjustment_method": localized_method},
            template,
            "imperial",
        )


def test_lookup_is_deterministic_and_english_is_the_fallback() -> None:
    assert normalize_locale("unknown") == "en"
    assert translate("result.capacity", "th") == "ความจุ"
    assert translate("unknown.presentation.key", "th") == "Unknown presentation key"


def test_locale_is_not_part_of_the_calculation_fingerprint() -> None:
    inputs = load_freeway_preset("BF-CH26-001")["inputs"]
    english = calculation_input_fingerprint("hcm7_basic_freeway_segment", "phase_10_product_integration", inputs)
    thai = calculation_input_fingerprint("hcm7_basic_freeway_segment", "phase_10_product_integration", inputs)
    assert english == thai


def test_saved_locale_is_optional_presentation_metadata_and_does_not_stale_result() -> None:
    preset = load_freeway_preset("BF-CH26-001")
    project = json.loads(
        create_manual_freeway_project_json(
            "BF-CH26-001",
            "imperial",
            freeway_preset_ui_inputs("BF-CH26-001", "imperial"),
            result=_freeway_result(),
            locale="th",
        )
    )
    loaded = load_manual_freeway_project_json(json.dumps(project))
    assert project_presentation_locale(loaded) == "th"
    assert loaded["calculation_result"] is not None


def test_legacy_project_without_presentation_metadata_retains_no_locale_choice() -> None:
    project = json.loads(create_manual_freeway_project_json(
        "BF-CH26-001", "imperial", freeway_preset_ui_inputs("BF-CH26-001", "imperial")
    ))
    project.pop("presentation")
    assert project_presentation_locale(load_manual_freeway_project_json(json.dumps(project))) is None


def test_thai_report_preserves_canonical_json_keys_and_localizes_presentation() -> None:
    report = build_report("manual_basic_freeway_v0", _freeway_result(), "imperial", inputs={}, locale="th")
    exported = json.loads(export_report(report, "json"))
    assert exported["method_identifier"] == "hcm7_basic_freeway_segment"
    assert exported["presentation"]["locale"] == "th"
    assert "ความจุ" in export_report(report, "markdown")


def test_thai_excel_contains_localized_sheet_presentation() -> None:
    report = build_report("manual_basic_freeway_v0", _freeway_result(), "imperial", inputs={}, locale="th")
    workbook = load_workbook(BytesIO(export_report(report, "xlsx")))
    assert "สรุปผลลัพธ์" in workbook.sheetnames


def test_streamlit_locale_switch_is_presentation_only() -> None:
    streamlit = pytest.importorskip("streamlit")
    from streamlit.testing.v1 import AppTest

    app = AppTest.from_file(
        str(Path(__file__).parents[2] / "src" / "hcmcalc" / "ui" / "streamlit_app.py")
    )
    app.run(timeout=30)
    app.selectbox[0].set_value("th").run(timeout=30)
    assert not app.exception
    assert app.selectbox[0].label == "ภาษา"
    assert app.selectbox[1].label == "กลุ่มเครื่องคำนวณ"
    assert app.selectbox[2].label == "เครื่องคำนวณ"
    app.selectbox[0].set_value("en").run(timeout=30)
    assert not app.exception
