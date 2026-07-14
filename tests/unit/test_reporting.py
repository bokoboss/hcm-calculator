import json
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from hcmcalc.cli import result_to_dict
from hcmcalc.ui.audit import build_manual_calculation_audit_record
from hcmcalc.ui.manual_facility import (
    build_manual_facility_audit_record,
    load_facility_template,
    run_manual_facility,
)
from hcmcalc.ui.manual_segment import run_manual_single_segment
from hcmcalc.ui.manual_multilane import (
    MANUAL_MULTILANE_LIMITATIONS,
    build_manual_multilane_audit_record,
    load_multilane_template,
    multilane_template_ui_inputs,
    multilane_ui_inputs_to_engine,
    run_manual_multilane,
)
from hcmcalc.ui.manual_freeway import (
    MANUAL_FREEWAY_LIMITATIONS,
    build_manual_freeway_audit_record,
    freeway_preset_ui_inputs,
    load_freeway_preset,
    run_manual_freeway,
)
from hcmcalc.ui.reporting import (
    FACILITY_LIMITATIONS,
    SINGLE_SEGMENT_LIMITATIONS,
    ReportingError,
    build_report,
    export_report,
    report_filename,
)
from hcmcalc.ui.units import manual_defaults


def _single_segment_context(unit_system: str = "metric"):
    inputs = {
        **manual_defaults(unit_system),
        "unit_system": unit_system,
        "segment_type": "passing_constrained",
        "terrain_type": "level",
        "horizontal_alignment": "straight",
    }
    result = run_manual_single_segment(inputs)
    return inputs, result_to_dict(result), build_manual_calculation_audit_record(
        inputs, result=result
    )


def _facility_context(unit_system: str = "metric"):
    template = load_facility_template("mountainous_example_4", unit_system)
    result = run_manual_facility(
        template["template_id"], template["segments"], unit_system
    )
    return (
        template,
        result_to_dict(result),
        build_manual_facility_audit_record(
            template["template_id"], template["segments"], unit_system, result=result
        ),
    )


def _single_report(unit_system: str = "metric"):
    inputs, result, audit = _single_segment_context(unit_system)
    return build_report(
        "manual_single_segment", result, unit_system, inputs=inputs, audit_record=audit
    )


def _facility_report(unit_system: str = "metric"):
    template, result, audit = _facility_context(unit_system)
    return build_report(
        "manual_two_lane_facility_v1",
        result,
        unit_system,
        inputs=template["segments"],
        audit_record=audit,
        template_id=template["template_id"],
    )


def _multilane_report(unit_system: str = "metric", template_id: str = "MLH-CH26-004-EB"):
    template = load_multilane_template(template_id)
    displayed = multilane_template_ui_inputs(template_id, unit_system)
    result = run_manual_multilane(template["inputs"])
    audit = build_manual_multilane_audit_record(
        template_id,
        template["inputs"],
        unit_system=unit_system,
        displayed_inputs=displayed,
        result=result,
    )
    return build_report(
        "manual_multilane_v0",
        result_to_dict(result),
        unit_system,
        inputs=displayed,
        audit_record=audit,
        template_id=template_id,
    )


def _freeway_report(unit_system: str = "metric", preset_id: str = "BF-CH26-001"):
    preset = load_freeway_preset(preset_id)
    displayed = freeway_preset_ui_inputs(preset_id, unit_system)
    result = run_manual_freeway(preset["inputs"])
    audit = build_manual_freeway_audit_record(
        preset_id,
        preset["inputs"],
        unit_system=unit_system,
        displayed_inputs=displayed,
        result=result,
    )
    return build_report(
        "manual_basic_freeway_v0",
        result_to_dict(result),
        unit_system,
        inputs=displayed,
        audit_record=audit,
        template_id=preset_id,
    )


def test_single_segment_report_json_generation_includes_units_and_limitations() -> None:
    report = _single_report()
    exported = json.loads(export_report(report, "json"))

    assert exported["schema_version"] == "0.1"
    assert exported["calculation_type"] == "manual_single_segment"
    assert any(item["unit"] == "km/h" for item in exported["results_summary"])
    assert exported["limitations"] == SINGLE_SEGMENT_LIMITATIONS
    assert exported["assumptions"]


def test_facility_report_json_generation_includes_segment_rows_and_limitations() -> None:
    report = _facility_report()
    exported = json.loads(export_report(report, "json"))

    assert exported["calculation_type"] == "manual_two_lane_facility_v1"
    assert len(exported["segment_results"]) == 6
    assert "Average speed (km/h)" in exported["segment_results"][0]
    assert exported["limitations"] == FACILITY_LIMITATIONS
    assert isinstance(exported["warnings"], list)


def test_multilane_report_json_includes_units_context_and_limitations() -> None:
    exported = json.loads(export_report(_multilane_report(), "json"))

    assert exported["calculation_type"] == "manual_multilane_v0"
    assert exported["selected_validated_template"] == "MLH-CH26-004-EB"
    assert any(item["unit"] == "km/h" for item in exported["results_summary"])
    assert any(
        item["unit"] == "ft (engine-native Imperial)"
        for item in exported["normalized_engine_inputs_summary"]
    )
    assert exported["limitations"] == MANUAL_MULTILANE_LIMITATIONS
    assert exported["assumptions"]
    assert exported["warnings"]
    assert exported["intermediate_values"]


def test_freeway_report_json_includes_units_context_results_and_limitations() -> None:
    exported = json.loads(export_report(_freeway_report(), "json"))

    assert exported["calculation_type"] == "manual_basic_freeway_v0"
    assert exported["selected_validated_preset"] == "BF-CH26-001"
    assert exported["support_scope"].startswith("Basic Freeway Segment only")
    assert any(item["unit"] == "km/h" for item in exported["results_summary"])
    assert any(
        item["label"] == "Adjusted Capacity" for item in exported["results_summary"]
    )
    assert any(
        item["unit"] == "mi (engine-native Imperial)"
        for item in exported["normalized_engine_inputs_summary"]
    )
    assert exported["limitations"] == MANUAL_FREEWAY_LIMITATIONS
    assert exported["assumptions"]
    assert exported["warnings"]
    assert exported["intermediate_values"]


@pytest.mark.parametrize(
    ("report_factory", "expected"),
    [
        (_single_report, "Manual Single Segment"),
        (_facility_report, "Manual Facility Report"),
            (_multilane_report, "Manual Multilane Highway Segment Report"),
        (_freeway_report, "Manual Basic Freeway Segment Calculator"),
    ],
)
def test_markdown_generation(report_factory, expected: str) -> None:
    markdown = export_report(report_factory(), "markdown")

    assert expected in markdown
    assert "## Limitations" in markdown
    assert "km/h" in markdown
    assert "## Assumptions" in markdown


@pytest.mark.parametrize(
    "report_factory",
    [_single_report, _facility_report, _multilane_report, _freeway_report],
)
def test_csv_generation_includes_units_limitations_and_warnings(report_factory) -> None:
    csv_text = export_report(report_factory(), "csv")

    assert "Limitations" in csv_text
    assert "Warnings" in csv_text
    assert "km/h" in csv_text
    assert "Assumptions" in csv_text


@pytest.mark.parametrize(
    "report_factory",
    [_single_report, _facility_report, _multilane_report, _freeway_report],
)
def test_excel_generation_has_required_sheets_units_and_limitations(report_factory) -> None:
    workbook = load_workbook(BytesIO(export_report(report_factory(), "xlsx")))

    assert workbook.sheetnames == [
        "Summary",
        "Inputs",
        "Segment Results",
        "Assumptions Warnings Limits",
        "Audit",
    ]
    cells = " ".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "km/h" in cells
    assert any(limitation in cells for limitation in report_factory()["limitations"])


def test_facility_csv_contains_segment_result_rows() -> None:
    csv_text = export_report(_facility_report("imperial"), "csv")

    assert "Segment Results" in csv_text
    assert "Segment ID,Segment type,Length (mi)" in csv_text
    assert csv_text.count("passing_constrained") >= 1


def test_multilane_metric_and_imperial_exports_use_selected_display_units() -> None:
    metric_report = _multilane_report("metric")
    imperial_report = _multilane_report("imperial")
    metric_csv = export_report(metric_report, "csv")
    imperial_csv = export_report(imperial_report, "csv")

    assert "pc/km/ln" in metric_csv
    assert "km/h" in metric_csv
    assert "pc/mi/ln" in imperial_csv
    assert "mph" in imperial_csv
    assert "Normalized Engine Inputs" in metric_csv
    assert "engine-native Imperial" in metric_csv
    assert next(
        item for item in imperial_report["inputs_summary"]
        if item["label"] == "Segment Length"
    )["unit"] == "ft"
    assert "Manual Multilane is a bounded" in metric_csv
    assert "## Intermediate Values" in export_report(
        _multilane_report("metric"), "markdown"
    )
    assert "## Normalized Engine Inputs" in export_report(
        _multilane_report("metric"), "markdown"
    )


def test_multilane_non_example_markdown_export_includes_key_outputs_and_scope() -> None:
    displayed = multilane_template_ui_inputs("MLH-CH26-004-EB", "imperial")
    displayed.update(
        {
            "number_of_lanes": 3,
            "segment_length": 5280.0,
            "demand_volume_veh_h": 2400.0,
            "peak_hour_factor": 0.92,
            "heavy_vehicle_percent": 12.0,
            "grade_percent": 0.0,
            "ffs_source": "measured",
            "free_flow_speed": 55.0,
            "passenger_car_equivalent": 2.0,
        }
    )
    inputs = load_multilane_template("MLH-CH26-004-EB")["inputs"] | {
        "number_of_lanes": 3,
        "segment_length_ft": 5280.0,
        "demand_volume_veh_h": 2400.0,
        "peak_hour_factor": 0.92,
        "heavy_vehicle_percent": 12.0,
        "grade_percent": 0.0,
        "ffs_source": "measured",
        "free_flow_speed_mph": 55.0,
        "passenger_car_equivalent": 2.0,
    }
    result = run_manual_multilane(inputs)
    audit = build_manual_multilane_audit_record(
        "MLH-CH26-004-EB",
        inputs,
        unit_system="imperial",
        displayed_inputs=displayed,
        result=result,
    )

    markdown = export_report(
        build_report(
            "manual_multilane_v0",
            result_to_dict(result),
            "imperial",
            inputs=displayed,
            audit_record=audit,
            template_id="MLH-CH26-004-EB",
        ),
        "markdown",
    )

    assert "Multilane Highway Segment" in markdown
    assert "Level of service" in markdown
    assert "Density" in markdown
    assert "Demand Flow Rate" in markdown
    assert "bounded_multilane_segment_phase_8" in markdown


def test_multilane_estimated_non_example_markdown_export_includes_key_outputs_and_scope() -> None:
    displayed = multilane_template_ui_inputs("MLH-CH26-004-EB", "imperial")
    displayed.update(
        {
            "segment_length": 5280.0,
            "demand_volume_veh_h": 1200.0,
            "peak_hour_factor": 0.95,
            "heavy_vehicle_percent": 15.0,
            "grade_percent": 0.0,
            "lane_width": 11.0,
            "roadside_lateral_clearance": 4.0,
            "access_point_density": 5.0,
            "ffs_source": "estimated",
            "passenger_car_equivalent": 2.5,
        }
    )
    inputs = load_multilane_template("MLH-CH26-004-EB")["inputs"] | {
        "segment_length_ft": 5280.0,
        "demand_volume_veh_h": 1200.0,
        "peak_hour_factor": 0.95,
        "heavy_vehicle_percent": 15.0,
        "grade_percent": 0.0,
        "lane_width_ft": 11.0,
        "roadside_lateral_clearance_ft": 4.0,
        "access_point_density_per_mi": 5.0,
        "ffs_source": "estimated",
        "passenger_car_equivalent": 2.5,
    }
    result = run_manual_multilane(inputs)
    audit = build_manual_multilane_audit_record(
        "MLH-CH26-004-EB",
        inputs,
        unit_system="imperial",
        displayed_inputs=displayed,
        result=result,
    )

    markdown = export_report(
        build_report(
            "manual_multilane_v0",
            result_to_dict(result),
            "imperial",
            inputs=displayed,
            audit_record=audit,
            template_id="MLH-CH26-004-EB",
        ),
        "markdown",
    )

    assert "Multilane Highway Segment" in markdown
    assert "Level of service" in markdown
    assert "Density" in markdown
    assert "Demand Flow Rate" in markdown
    assert "bounded_multilane_segment_phase_8" in markdown


def test_freeway_metric_and_imperial_exports_use_selected_display_units() -> None:
    metric_report = _freeway_report("metric")
    imperial_report = _freeway_report("imperial")
    metric_csv = export_report(metric_report, "csv")
    imperial_csv = export_report(imperial_report, "csv")
    markdown = export_report(metric_report, "markdown")

    assert "pc/km/ln" in metric_csv
    assert "km/h" in metric_csv
    assert "pc/mi/ln" in imperial_csv
    assert "mph" in imperial_csv
    assert "Normalized Engine Inputs" in metric_csv
    assert "engine-native Imperial" in metric_csv
    assert "Manual Basic Freeway Segment v0.1 supports bounded" in metric_csv
    assert "no ramps" in markdown
    assert "Adjusted Free Flow Speed" in markdown
    assert next(
        item for item in imperial_report["inputs_summary"]
        if item["label"] == "Segment Length"
    )["unit"] == "mi"


def test_freeway_non_example_markdown_export_includes_key_outputs_and_scope() -> None:
    displayed = freeway_preset_ui_inputs("BF-CH26-001", "imperial")
    displayed.update(
        {
            "number_of_lanes": 4,
            "segment_length": 2.6,
            "base_free_flow_speed": 74.0,
            "lane_width": 11.0,
            "right_side_lateral_clearance": 3.0,
            "total_ramp_density": 2.5,
            "demand_volume_veh_h": 5100.0,
            "peak_hour_factor": 0.95,
            "heavy_vehicle_percent": 4.0,
        }
    )
    inputs = load_freeway_preset("BF-CH26-001")["inputs"] | {
        "number_of_lanes": 4,
        "segment_length_mi": 2.6,
        "base_free_flow_speed_mph": 74.0,
        "lane_width_ft": 11.0,
        "right_side_lateral_clearance_ft": 3.0,
        "total_ramp_density_per_mi": 2.5,
        "demand_volume_veh_h": 5100.0,
        "peak_hour_factor": 0.95,
        "heavy_vehicle_percent": 4.0,
    }
    result = run_manual_freeway(inputs)
    audit = build_manual_freeway_audit_record(
        "BF-CH26-001",
        inputs,
        unit_system="imperial",
        displayed_inputs=displayed,
        result=result,
    )

    markdown = export_report(
        build_report(
            "manual_basic_freeway_v0",
            result_to_dict(result),
            "imperial",
            inputs=displayed,
            audit_record=audit,
            template_id="BF-CH26-001",
        ),
        "markdown",
    )

    assert "Basic Freeway Segment" in markdown
    assert "Level of service" in markdown
    assert "Density" in markdown
    assert "Demand Flow Rate" in markdown
    assert "supported_basic_freeway_segment_v0_1" in markdown


def test_report_generation_rejects_missing_result() -> None:
    with pytest.raises(ReportingError, match="calculated result is required"):
        build_report("manual_single_segment", None, "metric")


def test_report_generation_rejects_unsupported_calculation_type() -> None:
    with pytest.raises(ReportingError, match="Unsupported calculation_type"):
        build_report("validated_example", {"outputs": {}}, "metric")


def test_report_generation_rejects_malformed_result() -> None:
    with pytest.raises(ReportingError, match="Malformed result object"):
        build_report("manual_single_segment", {"outputs": {}}, "metric")


def test_report_export_rejects_unsupported_format() -> None:
    with pytest.raises(ReportingError, match="Unsupported export format"):
        export_report(_single_report(), "pdf")


def test_report_export_rejects_non_serializable_field() -> None:
    report = _single_report()
    report["inputs_summary"][0]["value"] = object()

    with pytest.raises(ReportingError, match="Non-serializable report field"):
        export_report(report, "json")


def test_reporting_module_is_streamlit_independent() -> None:
    source = Path("src/hcmcalc/ui/reporting.py").read_text(encoding="utf-8")

    assert "import streamlit" not in source


def test_report_filename_uses_required_pattern() -> None:
    report = _single_report()
    report["generated_at"] = "2026-06-12T10:11:12+00:00"

    assert (
        report_filename(report, "xlsx")
        == "hcm_ch15_single_segment_report_20260612_101112.xlsx"
    )


@pytest.mark.parametrize("calculation_type", ["manual_multilane_v0", "manual_basic_freeway_v0"])
def test_above_capacity_exports_preserve_absent_speed_and_density(
    calculation_type: str,
) -> None:
    if calculation_type == "manual_multilane_v0":
        template = load_multilane_template("MLH-CH26-004-EB")
        displayed = multilane_template_ui_inputs("MLH-CH26-004-EB", "imperial") | {
            "demand_volume_veh_h": 20000.0,
        }
        inputs = multilane_ui_inputs_to_engine(displayed, template["inputs"], "imperial")
        result = result_to_dict(run_manual_multilane(inputs))
        audit = build_manual_multilane_audit_record(
            "MLH-CH26-004-EB", inputs, displayed_inputs=displayed,
            result=run_manual_multilane(inputs),
        )
        template_id = "MLH-CH26-004-EB"
    else:
        preset = load_freeway_preset("BF-CH26-001")
        displayed = freeway_preset_ui_inputs("BF-CH26-001", "imperial") | {
            "demand_volume_veh_h": 20000.0,
        }
        inputs = preset["inputs"] | {"demand_volume_veh_h": 20000.0}
        result = result_to_dict(run_manual_freeway(inputs))
        audit = build_manual_freeway_audit_record(
            "BF-CH26-001", inputs, displayed_inputs=displayed,
            result=run_manual_freeway(inputs),
        )
        template_id = "BF-CH26-001"

    report = build_report(
        calculation_type, result, "imperial", inputs=displayed,
        audit_record=audit, template_id=template_id,
    )
    exported = json.loads(export_report(report, "json"))
    summary = {row["label"]: row["value"] for row in exported["results_summary"]}

    assert result["outputs"]["level_of_service"] == "F"
    assert summary["Speed Used For Density"] is None
    assert summary["Density"] is None
    assert "None" not in export_report(report, "markdown")
