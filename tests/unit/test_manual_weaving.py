import json
from io import BytesIO

import pytest
from openpyxl import load_workbook

from hcmcalc.cli import result_to_dict
from hcmcalc.core import HCMCalcError
from hcmcalc.ui.manual_weaving import (
    WEAVING_CALCULATION_CONTRACT,
    WEAVING_METHOD_VERSION,
    build_manual_weaving_audit_record,
    load_weaving_preset,
    run_manual_weaving,
    weaving_preset_ui_inputs,
    weaving_ui_inputs_to_engine,
)
from hcmcalc.ui.project_io import (
    MANUAL_WEAVING_PROJECT_TYPE,
    ProjectFileError,
    create_manual_weaving_project_json,
    load_manual_weaving_project_json,
)
from hcmcalc.ui.reporting import build_report, export_report
from hcmcalc.ui.workflow_state import calculation_input_fingerprint


@pytest.mark.parametrize("preset_id, expected_los", [
    ("WVG-CH27-001", "C"), ("WVG-CH27-002", "C"), ("WVG-CH27-003", "E"),
])
def test_reference_presets_map_only_to_the_qualified_public_engine(preset_id, expected_los):
    displayed = weaving_preset_ui_inputs(preset_id, "imperial")
    normalized = weaving_ui_inputs_to_engine(displayed, "imperial")
    result = run_manual_weaving(normalized)

    assert normalized["method_version"] == WEAVING_METHOD_VERSION
    assert result.outputs["level_of_service"] == expected_los


def test_metric_and_imperial_inputs_normalize_to_the_same_engine_case():
    imperial = weaving_preset_ui_inputs("WVG-CH27-001", "imperial")
    metric = weaving_preset_ui_inputs("WVG-CH27-001", "metric")

    assert weaving_ui_inputs_to_engine(metric, "metric") == weaving_ui_inputs_to_engine(imperial, "imperial")


def test_configuration_clears_inactive_lane_change_fields():
    displayed = weaving_preset_ui_inputs("WVG-CH27-003", "imperial")
    displayed["lc_rf"] = 1
    displayed["lc_fr"] = 1
    normalized = weaving_ui_inputs_to_engine(displayed, "imperial")

    assert normalized["lc_rf"] is None
    assert normalized["lc_fr"] is None
    assert normalized["lc_rr"] == 2


def test_hcm_7_1_is_not_an_adapter_or_project_execution_path():
    displayed = weaving_preset_ui_inputs("WVG-CH27-001", "imperial")
    normalized = weaving_ui_inputs_to_engine(displayed, "imperial")
    normalized["method_version"] = "hcm_7_1"
    with pytest.raises(HCMCalcError):
        run_manual_weaving(normalized)

    payload = json.loads(create_manual_weaving_project_json("WVG-CH27-001", "imperial", displayed))
    payload["method_version"] = "hcm_7_1"
    with pytest.raises(ProjectFileError, match="hcm_7_0"):
        load_manual_weaving_project_json(json.dumps(payload))


def _current_project_context(preset_id="WVG-CH27-001", unit_system="metric"):
    displayed = weaving_preset_ui_inputs(preset_id, unit_system)
    normalized = weaving_ui_inputs_to_engine(displayed, unit_system)
    result = run_manual_weaving(normalized)
    audit = build_manual_weaving_audit_record(
        preset_id, normalized, unit_system=unit_system, displayed_inputs=displayed, result=result
    )
    return displayed, normalized, result_to_dict(result), audit


def test_weaving_project_round_trip_preserves_identity_result_and_nulls():
    displayed, normalized, result, audit = _current_project_context()
    payload = json.loads(create_manual_weaving_project_json(
        "WVG-CH27-001", "metric", displayed, result=result, audit_record=audit
    ))
    loaded = load_manual_weaving_project_json(json.dumps(payload))

    assert payload["project_type"] == MANUAL_WEAVING_PROJECT_TYPE
    assert payload["method_version"] == WEAVING_METHOD_VERSION
    assert payload["calculation_contract"] == WEAVING_CALCULATION_CONTRACT
    assert loaded["load_status"] == "result_current"
    assert loaded["normalized_engine_inputs"] == normalized


def test_weaving_project_discards_stale_result_when_inputs_change():
    displayed, _, result, audit = _current_project_context()
    payload = json.loads(create_manual_weaving_project_json(
        "WVG-CH27-001", "metric", displayed, result=result, audit_record=audit
    ))
    payload["displayed_ui_inputs"]["volume_ff_veh_h"] += 10
    loaded = load_manual_weaving_project_json(json.dumps(payload))

    assert loaded["load_status"] == "project_requires_recalculation"
    assert loaded["calculation_result"] is None


@pytest.mark.parametrize("preset_id", ["WVG-CH27-001", "WVG-CH27-003"])
def test_all_exports_include_weaving_metadata_and_structurally_valid_excel(preset_id):
    displayed, normalized, result, audit = _current_project_context(preset_id, "metric")
    report = build_report(
        "manual_freeway_weaving_segment_v1", result, "metric", inputs=displayed,
        audit_record=audit, template_id=preset_id,
    )
    exported = json.loads(export_report(report, "json"))
    workbook = load_workbook(BytesIO(export_report(report, "xlsx")))

    assert exported["method_version"] == WEAVING_CALCULATION_CONTRACT
    assert exported["selected_validation_preset"] == preset_id
    assert any(row["unit"] == "km/h" for row in exported["results_summary"])
    assert workbook.sheetnames == ["Summary", "Inputs", "Segment Results", "Assumptions Warnings Limits", "Audit"]
    assert "weaving" in export_report(report, "markdown").lower()


def test_above_capacity_exports_null_predictions_and_handoff_exports_no_los():
    displayed, _, _, _ = _current_project_context(unit_system="imperial")
    displayed["volume_ff_veh_h"] = 25000.0
    normalized = weaving_ui_inputs_to_engine(displayed, "imperial")
    result = result_to_dict(run_manual_weaving(normalized))
    audit = build_manual_weaving_audit_record("WVG-CH27-001", normalized, unit_system="imperial", displayed_inputs=displayed, result=run_manual_weaving(normalized))
    report = build_report("manual_freeway_weaving_segment_v1", result, "imperial", inputs=displayed, audit_record=audit)
    summary = {row["label"]: row["value"] for row in report["results_summary"]}
    assert result["outputs"]["level_of_service"] == "F"
    assert summary["Mean speed"] is None and summary["Density"] is None

    displayed = weaving_preset_ui_inputs("WVG-CH27-001", "imperial")
    displayed["segment_length"] = 6000.0
    result = run_manual_weaving(weaving_ui_inputs_to_engine(displayed, "imperial"))
    assert result.outputs["support_status"] == "hcm_handoff_required"
    assert result.outputs["level_of_service"] is None


def test_fingerprint_covers_calculation_relevant_geometry_but_not_display_locale():
    displayed = weaving_preset_ui_inputs("WVG-CH27-001", "imperial")
    normalized = weaving_ui_inputs_to_engine(displayed, "imperial")
    baseline = calculation_input_fingerprint("weaving_segment", WEAVING_CALCULATION_CONTRACT, normalized)
    changed = {**normalized, "lc_fr": 2}
    assert calculation_input_fingerprint("weaving_segment", WEAVING_CALCULATION_CONTRACT, changed) != baseline
